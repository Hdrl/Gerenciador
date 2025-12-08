from viagens.models import TransacaoFinanceira
import zipfile
import tempfile
import os
import sys
import traceback
import datetime
from django.http import FileResponse
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from django.contrib import admin
from django.conf import settings

@admin.action(description='📒 Gerar Relatório de Viagem')
def gerar_relatorio_viagem(modeladmin, request, queryset):
    # --- BLINDAGEM 1: Try/Except Global para evitar Erro 500 genérico ---
    try:
        if len(queryset) > 1:
            return HttpResponse("Limitado a 1 relatório por vez. Por favor, selecione apenas uma viagem.")
        
        # Cria diretório temporário para evitar erros de permissão de escrita no servidor
        tmp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(tmp_dir, 'exportacao.zip')
        excel_output_path = os.path.join(tmp_dir, 'despesas.xlsx')
        
        viagem = queryset[0]
        despesas = TransacaoFinanceira.objects.filter(viagem=viagem, usuario=request.user, tipo='S').order_by('data')
        adiantamentos = TransacaoFinanceira.objects.filter(viagem=viagem, usuario=request.user, tipo='E').order_by('data')
        
        # --- BLINDAGEM 2: Caminho Absoluto para o Template ---
        # Tenta achar o arquivo usando o BASE_DIR do Django para funcionar em qualquer servidor
        caminho_template = os.path.join(settings.BASE_DIR, 'relatorio', 'relatorio_despesas.xlsx')
        
        # Fallback: Se não achar pelo BASE_DIR, tenta relativo (ambiente local)
        if not os.path.exists(caminho_template):
            caminho_template = 'relatorio/relatorio_despesas.xlsx'
            
        if not os.path.exists(caminho_template):
             return HttpResponse(f"Erro Crítico: O template '{caminho_template}' não foi encontrado no servidor.", status=500)

        wb = load_workbook(caminho_template)
        ws = wb['Despesas']

        # Preenchimento do Cabeçalho
        ws['A2'] = f"Empresa: {viagem.empresa}" 
        ws['C2'] = f"Setor: {viagem.setor}"
        ws['A3'] = f"Colaborador: {viagem.colaborador}"
        ws['C3'] = f"Retorno: {viagem.retorno.strftime('%d/%m/%Y') if viagem.retorno else ''}"
        ws['A4'] = f"Destino: {viagem.destino}"
        ws['B4'] = f"Motivo Viagem: {viagem.motivo}"
        ws['A5'] = f"Saída: {viagem.saida}"
        
        # Lista para registrar arquivos que faltaram no disco
        arquivos_com_erro = []

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            # Estilos
            borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fonte_hyperlink = Font(color='0563C1', underline='single', name='Arial', size=12)
            fonte_titulo = Font(bold=True, name='Arial', size=12)
            fonte_normal = Font(name='Arial', size=12)
            fonte_erro = Font(color='FF0000', bold=True, name='Arial', size=10) # Vermelho para erros
            alinhamento = Alignment(horizontal='center', vertical='center')

            # Formatação Geral
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alinhamento
                    if 1 <= cell.row <= 8 and 1 <= cell.column <= 4:
                        cell.font = fonte_titulo
                    else:
                        cell.font = fonte_normal
                
            # --- LOOP DESPESAS ---
            for i, despesa in enumerate(despesas, start=9):
                descricao_despesa = ws.cell(row=i, column=1)
                data_despesa = ws.cell(row=i, column=2)
                nota_despesa = ws.cell(row=i, column=3)
                valor_despesa = ws.cell(row=i, column=4)
                
                if despesa.descricao: descricao_despesa.value = despesa.descricao.lower()
                if despesa.data: data_despesa.value = despesa.data.strftime('%d/%m/%Y %H:%M')
                if despesa.valor: valor_despesa.value = despesa.valor
                valor_despesa.number_format = 'R$ #,##0.00'
                
                # --- BLINDAGEM 3: Verificação de Existência de Imagem ---
                if despesa.imagem:
                    try:
                        caminho_real = despesa.imagem.path
                        if os.path.exists(caminho_real):
                            nome_arquivo = os.path.basename(caminho_real)
                            nota_despesa.value = f"=hyperlink(\"notas fiscal/{nome_arquivo}\", \"{nome_arquivo}\")"
                            nota_despesa.font = fonte_hyperlink
                            zipf.write(caminho_real, arcname=f"notas fiscal/{nome_arquivo}")
                        else:
                            # Se o arquivo não existir fisicamente
                            nota_despesa.value = "ARQUIVO NÃO ENCONTRADO"
                            nota_despesa.font = fonte_erro
                            arquivos_com_erro.append(f"Despesa ID {despesa.id}: {caminho_real}")
                    except Exception as e:
                        # Se der erro de permissão ou caminho inválido
                        nota_despesa.value = "ERRO NO ARQUIVO"
                        nota_despesa.font = fonte_erro
                        arquivos_com_erro.append(f"Despesa ID {despesa.id} (Erro Código): {str(e)}")

                descricao_despesa.border = borda
                data_despesa.border = borda
                nota_despesa.border = borda
                valor_despesa.border = borda
            
            # Totais
            total_row = len(despesas) + 9
            total_cell = ws.cell(row=total_row, column=1)
            total_cell.value = "TOTAL"
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
            total_cell.border = borda
            ws.cell(row=total_row, column=2).border = borda
            ws.cell(row=total_row, column=3).border = borda
            total_cell.font = fonte_titulo
            
            soma_cell = ws.cell(row=total_row, column=4)
            soma_cell.value = f"=sum(D9:D{len(despesas) + 9 - 1})"  
            soma_cell.border = borda
            
            # Cabeçalho Adiantamentos
            cabecalho_row = total_row + 2
            cabecalho_adiantamento = ws.cell(row=cabecalho_row, column=1)
            cabecalho_adiantamento.value = "Descricão Do Adiantamento"
            ws.merge_cells(start_row = cabecalho_row, start_column = 1, end_row=cabecalho_row, end_column=4)
            cabecalho_adiantamento.font = fonte_titulo
            
            descricao_adiantamento_lbl = ws.cell(row=cabecalho_row + 1, column=1)
            data_adiantamento_lbl = ws.cell(row=cabecalho_row + 1, column=2)
            comprovante_adiantamento_lbl = ws.cell(row=cabecalho_row + 1, column=3)
            valor_adiantamento_lbl = ws.cell(row=cabecalho_row + 1, column=4)
            
            descricao_adiantamento_lbl.value = "Descrição"
            data_adiantamento_lbl.value = "Data"
            comprovante_adiantamento_lbl.value = "Comprovante"
            valor_adiantamento_lbl.value = "Valor"
            
            # Bordas cabeçalho adiantamento
            cabecalho_adiantamento.border = borda
            ws.cell(row=cabecalho_row, column=2).border = borda
            ws.cell(row=cabecalho_row, column=3).border = borda
            ws.cell(row=cabecalho_row, column=4).border = borda
            descricao_adiantamento_lbl.border = borda
            data_adiantamento_lbl.border = borda
            comprovante_adiantamento_lbl.border = borda
            valor_adiantamento_lbl.border = borda
            
            # --- LOOP ADIANTAMENTOS ---
            adiantamento_row = cabecalho_adiantamento.row + 1
            for i, adiantamento in enumerate(adiantamentos, cabecalho_adiantamento.row + 2):
                descricao_adiantamento = ws.cell(row=i, column=1)
                data_adiantamento = ws.cell(row=i, column=2)
                comprovante_adiantamento = ws.cell(row=i, column=3)
                valor_adiantamento = ws.cell(row=i, column=4)
                
                if adiantamento.descricao: descricao_adiantamento.value = adiantamento.descricao.lower()
                if adiantamento.data: data_adiantamento.value = adiantamento.data.strftime('%d/%m/%Y %H:%M')
                if adiantamento.valor: valor_adiantamento.value = adiantamento.valor
                valor_adiantamento.number_format = 'R$ #,##0.00'
                
                # --- BLINDAGEM 3 (Repetição para Adiantamentos) ---
                if adiantamento.imagem:
                    try:
                        caminho_real = adiantamento.imagem.path
                        if os.path.exists(caminho_real):
                            nome_arquivo = os.path.basename(caminho_real)
                            comprovante_adiantamento.value = f"=hyperlink(\"notas fiscal/{nome_arquivo}\", \"{nome_arquivo}\")"
                            comprovante_adiantamento.font = fonte_hyperlink
                            zipf.write(caminho_real, arcname=f"notas fiscal/{nome_arquivo}")
                        else:
                            comprovante_adiantamento.value = "ARQUIVO NÃO ENCONTRADO"
                            comprovante_adiantamento.font = fonte_erro
                            arquivos_com_erro.append(f"Adiantamento ID {adiantamento.id}: {caminho_real}")
                    except Exception as e:
                        comprovante_adiantamento.value = "ERRO NO ARQUIVO"
                        comprovante_adiantamento.font = fonte_erro
                        arquivos_com_erro.append(f"Adiantamento ID {adiantamento.id}: {str(e)}")
                
                descricao_adiantamento.border = borda
                data_adiantamento.border = borda
                comprovante_adiantamento.border = borda
                valor_adiantamento.border = borda
                adiantamento_row = i # Atualiza a linha final
                
            # Totais Finais
            receber = ws.cell(row=adiantamento_row+2, column=1)
            devolver = ws.cell(row=adiantamento_row+2, column=3)
            valor_receber = ws.cell(row=adiantamento_row+3, column=1)
            valor_devolver = ws.cell(row=adiantamento_row+3, column=3)
            
            ws.merge_cells(start_row=receber.row, end_row=receber.row, start_column=1, end_column=2)
            ws.merge_cells(start_row=devolver.row, end_row=devolver.row, start_column=3, end_column=4)
            ws.merge_cells(start_row=receber.row+1, end_row=receber.row+1, start_column=1, end_column=2)
            ws.merge_cells(start_row=devolver.row+1, end_row=devolver.row+1, start_column=3, end_column=4)
            
            ws.cell(row=devolver.row, column=4).border = borda
            ws.cell(row=valor_devolver.row, column=4).border = borda
            ws.cell(row=receber.row, column=2).border = borda
            ws.cell(row=valor_receber.row, column=2).border = borda
            receber.border = borda
            devolver.border = borda
            valor_receber.border = borda
            valor_devolver.border = borda
            
            receber.value = "Receber"
            devolver.value = "Devolver"
            receber.font = fonte_titulo
            devolver.font = fonte_titulo
            
            valor_receber.value = f"=if(sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) - D{soma_cell.row} <0, D{soma_cell.row} - sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}), 0)"
            valor_devolver.value = f"=if(sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) - D{soma_cell.row}<0, 0, sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) -  D{soma_cell.row})"
            valor_devolver.number_format = 'R$ #,##0.00'
            valor_receber.number_format = 'R$ #,##0.00'
            
            # Rodapé
            gerado_em = ws.cell(row=valor_devolver.row + 2, column=1)
            ws.merge_cells(start_row=gerado_em.row, end_row=gerado_em.row, start_column=1, end_column=4)
            gerado_em.value = f"Apucarana, {datetime.datetime.now().strftime('%d de %B de %Y')} "
            
            ass_colaborador = ws.cell(row=gerado_em.row + 2, column=1)
            ass_financeiro = ws.cell(row=gerado_em.row + 2, column=3)
            valor_ass_colaborador = ws.cell(row=ass_colaborador.row+1, column=1)
            valor_ass_financeiro = ws.cell(row=ass_financeiro.row+1, column=3)
            
            ws.merge_cells(start_row=ass_colaborador.row, end_row=ass_colaborador.row, start_column=1, end_column=2)
            ws.merge_cells(start_row=ass_financeiro.row, end_row=ass_financeiro.row, start_column=3, end_column=4)
            ws.merge_cells(start_row=valor_ass_colaborador.row, end_row=valor_ass_colaborador.row, start_column=1, end_column=2)
            ws.merge_cells(start_row=valor_ass_financeiro.row, end_row=valor_ass_financeiro.row, start_column=3, end_column=4)
            
            ass_colaborador.value = "__________________________________"
            ass_financeiro.value = "__________________________________"
            valor_ass_colaborador.value = "Ass. Colaborador"
            valor_ass_financeiro.value = "Ass. Financeiro"       
            
            # --- SALVAMENTO SEGURO ---
            # Salva no diretório temporário para evitar erro de permissão
            wb.save(excel_output_path)
            zipf.write(excel_output_path, arcname="despesas.xlsx")

            # --- RELATÓRIO DE ERROS ---
            # Se houver arquivos perdidos, cria um TXT dentro do ZIP
            if arquivos_com_erro:
                txt_erro_path = os.path.join(tmp_dir, 'LEIA-ME_ERROS.txt')
                with open(txt_erro_path, 'w', encoding='utf-8') as f:
                    f.write("RELATÓRIO DE ARQUIVOS NÃO ENCONTRADOS\n")
                    f.write("=====================================\n\n")
                    f.write("O sistema não encontrou os seguintes arquivos de imagem no servidor:\n")
                    for erro in arquivos_com_erro:
                        f.write(f"- {erro}\n")
                    f.write("\nPossível causa: Os arquivos não foram enviados para a pasta 'media' do servidor.")
                
                zipf.write(txt_erro_path, arcname="LEIA-ME_ERROS.txt")

        # Retorna o arquivo ZIP final
        return FileResponse(open(zip_path, 'rb'), as_attachment=True, filename='despesas.zip')

    except Exception as e:
        # Imprime o erro no log do servidor (para o programador ver no journalctl)
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!", file=sys.stderr)
        
        # Retorna mensagem amigável para o usuário (navegador)
        return HttpResponse(f"<h1>Ocorreu um erro ao gerar o relatório</h1><p>Detalhe técnico: {str(e)}</p><p>Verifique se todas as imagens existem ou contate o suporte.</p>", status=500)